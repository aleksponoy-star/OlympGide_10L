from openpyxl import load_workbook
wb = load_workbook('org.xlsx')
ws = wb.active
import csv
with open('olympiad_organizers.csv', 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['olympiad_name', 'organizer_name'])
    for row in ws.iter_rows(values_only=True):
        name=row[0]
        org=row[1]
        if name==None and org==None:
            continue
        else:
            if name!=None:
                name1=name
            else:
                name=name1
        writer.writerow([name, org])