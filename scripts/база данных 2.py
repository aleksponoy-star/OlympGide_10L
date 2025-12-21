from openpyxl import load_workbook
import csv

wb = load_workbook('napr.xlsx')
ws = wb.active

with open('olympiad_napr.csv', 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['olympiad_name', 'profil_name', 'predmet_name', 'level_name'])

    name1 = None

    for row in ws.iter_rows(values_only=True):
        name = row[0]
        prof = row[1]
        pred = row[2]
        level = row[3]

        if name is None and prof is None and pred is None and level is None:
            continue

        if name is not None:
            name1 = name
        else:
            name = name1

        prof_list = [prof]
        if prof is not None and ',' in str(prof):
            prof_list = [p.strip() for p in str(prof).replace('\n', ' ').split(',')]

        pred_list = [pred]
        if pred is not None and ',' in str(pred):
            pred_list = [p.strip() for p in str(pred).replace('\n', ' ').split(',')]

        for prof_item in prof_list:
            for pred_item in pred_list:
                writer.writerow([name, prof_item, pred_item, level])

        
